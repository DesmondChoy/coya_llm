import sqlite3
import openpyxl
from contextlib import closing
import os

DB_PATH = "data.db"


def create_table_and_import_data(cursor, sheet):
    table_name = sheet.title
    headers = [cell.value for cell in sheet[1]]

    print(f"\n=== Importing Data for {table_name} ===")
    print("Headers:", headers)

    # Create table
    columns = ", ".join(f'"{header}" TEXT' for header in headers)
    cursor.execute(f"""
    CREATE TABLE IF NOT EXISTS "{table_name}" (
        id INTEGER PRIMARY KEY,
        {columns}
    )
    """)

    # Debug: Print Excel data before import
    print(f"\nData from Excel sheet {table_name}:")
    for row in sheet.iter_rows(min_row=2, values_only=True):
        print(row)

    # Import data
    insert_query = f'INSERT INTO "{table_name}" ({", ".join(headers)}) VALUES ({", ".join("?" * len(headers))})'
    cursor.executemany(insert_query, sheet.iter_rows(min_row=2, values_only=True))


def print_table_contents(cursor, table_name):
    cursor.execute(f'SELECT * FROM "{table_name}"')
    rows = cursor.fetchall()

    if rows:
        headers = [description[0] for description in cursor.description]
        print(f"\nContents of table '{table_name}':")
        print(", ".join(headers))
        for row in rows:
            print(", ".join(str(value) for value in row))
    else:
        print(f"\nTable '{table_name}' is empty.")


def create_database():
    with closing(sqlite3.connect(DB_PATH)) as conn, closing(conn.cursor()) as cursor:
        with closing(openpyxl.load_workbook("coya_app.xlsx")) as workbook:
            for sheet in workbook.worksheets:
                create_table_and_import_data(cursor, sheet)

        conn.commit()

        # Print contents of each table
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
        tables = cursor.fetchall()
        for (table_name,) in tables:
            print_table_contents(cursor, table_name)

    print("\nDatabase created and data imported successfully.")


def check_and_create_database():
    if not os.path.exists(DB_PATH):
        create_database()


def get_topics():
    with sqlite3.connect(DB_PATH) as conn:
        cursor = conn.cursor()

        cursor.execute("SELECT DISTINCT Topic FROM Story")
        story_topics = [row[0] for row in cursor.fetchall()]

        cursor.execute("SELECT DISTINCT Topic FROM Learn")
        learn_topics = [row[0] for row in cursor.fetchall()]

    return story_topics, learn_topics


def get_story_elements(topic):
    with sqlite3.connect(DB_PATH) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT Element, Description FROM Story WHERE Topic = ?", (topic,)
        )
        results = cursor.fetchall()
        print("\n=== Raw Database Query Results ===")
        print(f"Topic: {topic}")
        print("Results from database:")
        for row in results:
            print(f"Element: {row[0]}, Description: {row[1]}")
        print("=" * 50)
        return results


def get_quiz_data(topic):
    with sqlite3.connect(DB_PATH) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT Question, Answer, Truth FROM Learn WHERE Topic = ?", (topic,)
        )
        return cursor.fetchall()


def debug_story_and_quiz(topic):
    print(f"\n=== Debug Output for Topic: {topic} ===")

    print("\n--- Story Elements ---")
    story_elements = get_story_elements(topic)
    for element, description in story_elements:
        print(f"Element: {element}")
        print(f"Description: {description}")
        print("-" * 50)

    print("\n--- Quiz Data ---")
    quiz_data = get_quiz_data(topic)
    for question, answer, truth in quiz_data:
        print(f"Question: {question}")
        print(f"Answer: {answer}")
        print(f"Truth: {truth}")
        print("-" * 50)


if __name__ == "__main__":
    check_and_create_database()

    # Debug example - uncomment and modify topic to test
    story_topics, learn_topics = get_topics()
    if story_topics:
        print("\nAvailable Topics:", story_topics)
        debug_story_and_quiz(story_topics[0])  # Test with first available topic
